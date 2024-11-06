import openpyxl
import os
import pickle
import requests
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Clients_msg, News
from datetime import date, timedelta
from .forms import CreateUserForm, UpdateUserForm



# Create your views here.

def home(request):
    if request.method == 'POST':
        if 'Clients_msgs_form' in request.POST:
            name_contact_us = request.POST.get('name_contact_us','')
            email_contact_us = request.POST.get('email_contact_us','')
            message_contact_us = request.POST.get('message_contact_us','')
            if message_contact_us == '' or name_contact_us == '' or email_contact_us == '':
                return render(request, 'pages/home.html')
            else:
                data = Clients_msg(name_contact_us=name_contact_us,email_contact_us=email_contact_us,message_contact_us=message_contact_us)
                data.save()
        
    return render(request,'pages/home.html',{'news_fields':News.objects.all()})

def news(request):
    return render(request,'pages/news.html', {'news_fields':News.objects.all()})

def matches(request):
    data_matches = openpyxl.load_workbook('data_matches.xlsx')
    sheet = data_matches['features']
    selected_option = "bydate"
    selected_round = None
    if request.method == 'POST': 
        selected_option = request.POST.get('select-option')
        selected_round = request.POST.get('select-round')
        if selected_option is None and selected_round is None:
            selected_option = "bydate"
    today = date.today()
    seven_days_after = today + timedelta(days=7)
    
    team_images = {
        'Real Madrid' : 'team_images/real-madrid.png',  
        'FC Barcelona' : 'team_images/fc-barcelona.png',
        'CA Osasuna' : 'team_images/osasuna.png',
        'RC Celta' : 'team_images/celta.png',
        'Real Valladolid CF' : 'team_images/valladolid.png',
        'Cádiz CF' : 'team_images/cadiz.png',  
        'Valencia CF' : 'team_images/valencia.png',  
        'UD Almería' : 'team_images/almeria.png',  
        'Athletic Club' : 'team_images/athletic.png',  
        'Getafe CF' : 'team_images/getafe.png',  
        'Real Betis' : 'team_images/real-betis.png',  
        'Sevilla FC' : 'team_images/sevilla.png',  
        'RCD Espanyol' : 'team_images/espanyol.png',  
        'Villarreal CF' : 'team_images/villarreal.png',  
        'Rayo Vallecano' : 'team_images/rayo.png',  
        'Real Sociedad' : 'team_images/real-sociedad.png',  
        'Girona FC' : 'team_images/girona.png',  
        'RCD Mallorca' : 'team_images/mallorca.png',  
        'Atlético de Madrid' : 'team_images/atletico-de-madrid.png',  
        'Elche CF' : 'team_images/elche.png',  
    }

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        match_day = row[0]
        home_team = row[9]
        away_team = row[12]
        match_time = row[3]
        stadium =  row[8]
        match_date = row[1].date()
        match_month_day = (match_date.month, match_date.day)
        today_month_day = (today.month, today.day)
        seven_days_after_month_day = (seven_days_after.month, seven_days_after.day)
        
        if selected_option == home_team or selected_option == away_team:
            if match_month_day >= today_month_day:
                data.append({
                    'home_team': home_team,
                    'away_team': away_team,
                    'match_time': match_time,
                    'stadium': stadium,
                    'home_team_images': team_images[home_team],
                    'away_team_images': team_images[away_team],
                # add more fields as needed
                })
        elif selected_round is not None:
            selected_round = int(selected_round)
            if selected_round == match_day:
                data.append({
                        'home_team': home_team,
                        'away_team': away_team,
                        'match_time': match_time,
                        'stadium': stadium,
                        'home_team_images': team_images[home_team],
                        'away_team_images': team_images[away_team],
                    # add more fields as needed
                    })
        elif selected_option == "bydate":
            if match_month_day > seven_days_after_month_day:
                break
            elif match_month_day >= today_month_day:
                data.append({
                    'home_team': home_team,
                    'away_team': away_team,
                    'match_time': match_time,
                    'stadium': stadium,
                    'home_team_images': team_images[home_team],
                    'away_team_images': team_images[away_team],
                # add more fields as needed
                })
    return render(request,'pages/matches.html',{'data': data})

def scores(request):
    data_matches = openpyxl.load_workbook('data_matches.xlsx')
    sheet = data_matches['features']
    selected_option = "bylastround"
    selected_round = None
    if request.method == 'POST': 
        selected_option = request.POST.get('select-option')
        selected_round = request.POST.get('select-round')
        if selected_option is None and selected_round is None:
            selected_option = "bylastround"
    today = date.today()
    seven_days_after = today + timedelta(days=7)
    
    team_images = {
        'Real Madrid' : 'team_images/real-madrid.png',  
        'FC Barcelona' : 'team_images/fc-barcelona.png',
        'CA Osasuna' : 'team_images/osasuna.png',
        'RC Celta' : 'team_images/celta.png',
        'Real Valladolid CF' : 'team_images/valladolid.png',
        'Cádiz CF' : 'team_images/cadiz.png',  
        'Valencia CF' : 'team_images/valencia.png',  
        'UD Almería' : 'team_images/almeria.png',  
        'Athletic Club' : 'team_images/athletic.png',  
        'Getafe CF' : 'team_images/getafe.png',  
        'Real Betis' : 'team_images/real-betis.png',  
        'Sevilla FC' : 'team_images/sevilla.png',  
        'RCD Espanyol' : 'team_images/espanyol.png',  
        'Villarreal CF' : 'team_images/villarreal.png',  
        'Rayo Vallecano' : 'team_images/rayo.png',  
        'Real Sociedad' : 'team_images/real-sociedad.png',  
        'Girona FC' : 'team_images/girona.png',  
        'RCD Mallorca' : 'team_images/mallorca.png',  
        'Atlético de Madrid' : 'team_images/atletico-de-madrid.png',  
        'Elche CF' : 'team_images/elche.png',  
    }

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        match_day = row[0]
        home_team = row[9]
        away_team = row[12]
        match_time = row[3]
        stadium =  row[8]
        home_team_score = row[10]
        away_team_score = row[11]
        match_date = row[2].date()
        if selected_option == home_team or selected_option == away_team:
            if home_team_score is not None:
                data.append({
                    'home_team': home_team,
                    'away_team': away_team,
                    'match_time': match_time,
                    'stadium': stadium,
                    'home_team_score':home_team_score,
                    'away_team_score':away_team_score,
                    'home_team_images': team_images[home_team],
                    'away_team_images': team_images[away_team],
                # add more fields as needed
                })
        elif selected_round is not None:
            selected_round = int(selected_round)
            if selected_round == match_day:
                data.append({
                        'home_team': home_team,
                        'away_team': away_team,
                        'match_time': match_time,
                        'stadium': stadium,
                        'home_team_score':home_team_score,
                        'away_team_score':away_team_score,
                        'home_team_images': team_images[home_team],
                        'away_team_images': team_images[away_team],
                    # add more fields as needed
                    })
        elif selected_option == "bylastround":
            if match_day >= 29:
                break
            elif match_day == 28:
                data.append({
                    'home_team': home_team,
                    'away_team': away_team,
                    'match_time': match_time,
                    'stadium': stadium,
                    'home_team_score':home_team_score,
                    'away_team_score':away_team_score,
                    'home_team_images': team_images[home_team],
                    'away_team_images': team_images[away_team],
                # add more fields as needed
                })
    return render(request, 'pages/scores.html', {'data':data})

def standings(request):
    Daily_data = openpyxl.load_workbook('Daily_data.xlsx')
    sheet = Daily_data['Classement']
    team_images = {
        'Real Madrid ' : 'team_images/real-madrid.png',  
        'FC Barcelone ' : 'team_images/fc-barcelona.png',
        'Osasuna ' : 'team_images/osasuna.png',
        'Celta Vigo ' : 'team_images/celta.png',
        'Valladolid ' : 'team_images/valladolid.png',
        'Cadix ' : 'team_images/cadiz.png',  
        'Valence CF ' : 'team_images/valencia.png',  
        'Almeria ' : 'team_images/almeria.png',  
        'Athletic Bilbao ' : 'team_images/athletic.png',  
        'Getafe ' : 'team_images/getafe.png',  
        'Betis Séville ' : 'team_images/real-betis.png',  
        'Séville FC ' : 'team_images/sevilla.png',  
        'Espanyol Barcelone ' : 'team_images/espanyol.png',  
        'Villarreal ' : 'team_images/villarreal.png',  
        'Rayo Vallecano ' : 'team_images/rayo.png',  
        'Real Sociedad ' : 'team_images/real-sociedad.png',  
        'Gérone ' : 'team_images/girona.png',  
        'Majorque ' : 'team_images/mallorca.png',  
        'Atlético de Madrid ' : 'team_images/atletico-de-madrid.png',  
        'Elche ' : 'team_images/elche.png',  
    }

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        teams = row[1]
        pts = row[3]
        gp = row[4]
        w = row[5]
        d = row[6]
        l = row[7]
        f = row[8]
        a = row[9]
        gd = row[10]

        data.append({
            'teams':teams,
            'pts':pts,
            'gp':gp,
            'w':w,
            'd':d,
            'l':l,
            'f':f,
            'a':a,
            'gd':gd,
            'team_image': team_images[teams]
        })

    return render(request, 'pages/standings.html',{'data':data})

def teamstats(request):
    Teams_Stats = openpyxl.load_workbook('Teams_Stats.xlsx')
    sheet = Teams_Stats['team']
    team_images = {
        'Real Madrid' : 'team_images/real-madrid.png',  
        'Barcelona' : 'team_images/fc-barcelona.png',
        'Osasuna' : 'team_images/osasuna.png',
        'Celta Vigo' : 'team_images/celta.png',
        'Valladolid' : 'team_images/valladolid.png',
        'Cádiz' : 'team_images/cadiz.png',  
        'Valencia' : 'team_images/valencia.png',  
        'Almería' : 'team_images/almeria.png',  
        'Athletic Club' : 'team_images/athletic.png',  
        'Getafe' : 'team_images/getafe.png',  
        'Betis' : 'team_images/real-betis.png',  
        'Sevilla' : 'team_images/sevilla.png',  
        'Espanyol' : 'team_images/espanyol.png',  
        'Villarreal' : 'team_images/villarreal.png',  
        'Rayo Vallecano' : 'team_images/rayo.png',  
        'Real Sociedad' : 'team_images/real-sociedad.png',  
        'Girona' : 'team_images/girona.png',  
        'Mallorca' : 'team_images/mallorca.png',  
        'Atlético Madrid' : 'team_images/atletico-de-madrid.png',  
        'Elche' : 'team_images/elche.png',  
    }

    data = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        team_name = row[0]
        JC = row[1]
        Age = row[2]
        Poss =  row[3]
        Mj =  row[4]
        Titulaire =  row[5]
        Min =  row[6]
        Mindivby90 =  row[7]
        Buts =  row[8]
        PD =  row[9]
        BPD =  row[10]
        BPenM =  row[11]
        PenM =  row[12]
        PenT =  row[13]
        CJ =  row[14]
        CR =  row[15]
        xG =  row[16]
        npxG =  row[17]
        xAG =  row[18]
        npxG_xAG =  row[19]
        PrgC =  row[20]
        PrgP =  row[21]
        Butby90 =  row[22]
        PDby90 =  row[23]
        BPDby90 =  row[24]
        BPenMby90 =  row[25]
        BPDPenMby90 =  row[26]
        data.append({
            'team_name': team_name,
            'JC': JC,
            'Age':Age,
            'Poss': Poss,
            'Mj':Mj,
            'Titulaire': Titulaire,
            'Min': Min,
            'Mindivby90':Mindivby90,
            'Buts': Buts,
            'PD':PD,
            'BPD':BPD,
            'BPenM':BPenM,
            'PenM':PenM,
            'PenT':PenT,
            'CJ':CJ,
            'CR':CR,
            'xG':xG,
            'npxG':npxG,
            'xAG':xAG,
            'npxG_xAG':npxG_xAG,
            'PrgC':PrgC,
            'PrgP':PrgP,
            'Butby90':Butby90,
            'PDby90':PDby90,
            'BPDby90':BPDby90,
            'BPenMby90':BPenMby90,
            'BPDPenMby90':BPDPenMby90,
            'team_image':team_images[team_name]
        })
    return render(request, 'pages/teamstats.html',{'data': data})

def playerstats(request):
    response = requests.get('https://cdn.jsdelivr.net/npm/country-flag-emoji-json@2.0.0/dist/index.json')
    json_data = response.json()

    Stats = openpyxl.load_workbook('Stats.xlsx')
    nations = openpyxl.load_workbook('nations.xlsx')
    sheet = Stats['Stats']
    sheetN = nations['nations']
    selected_option = 'Real Madrid'
    if request.method == 'POST': 
        selected_option = request.POST.get('select-option')
        if selected_option is None:
            selected_option = 'Real Madrid'
    
    

    data = []
    for row in sheet.iter_rows(min_row=3,values_only=True):
         Joueur = row[1]
         Nation = row[2]
         Nation_name = None
         Nation_image = None
         Pos = row[3]
         Equipe = row[4]
         Age = row[5]
         Naissance = row[6]
         MJ = row[7]
         Titulaire = row[8]
         Min = row[9]
         Mindivby90 = row[10]
         Buts = row[11]
         PD = row[12]
         B_PD = row[13]
         B_PenM = row[14]
         PenM = row[15]
         PenT = row[16]
         CJ = row[17]
         CR = row[18]
         xG = row[19]
         npxG = row[20]
         xAG = row[21]
         npxG_xAG = row[22]
         PrgC = row[23]
         PrgP = row[24]
         PrgR = row[25]
         Butsby90 = row[26]
         PDby90 = row[27]
         B_PDby90 = row[28]
         BPenMby90 = row[29]
         B_PD_PenMby90 = row[30]
         nation_splited = Nation.split()[1]
         optionneeded = Equipe
         for row in sheetN.iter_rows(min_row=2,values_only=True):
             nations_name = row[1]
             nations_abrv = row[0]
             if nations_abrv == nation_splited:
                 Nation_name = nations_name
                 print(Nation_name)
                 break
         if Nation_name is not None:
             for country in json_data:
                 if country["name"] == Nation_name:
                     Nation_image = country["image"]
                     print(Nation_image)
                     break
         if selected_option != optionneeded:
             continue
         else:
             data.append({
                'Joueur':Joueur,
                'nation_splited':nation_splited,
                'Nation_name':Nation_name,
                'Nation_image':Nation_image,
                'Pos':Pos,
                'Equipe':Equipe,
                'Age':Age,
                'Naissance':Naissance,
                'MJ':MJ,
                'Titulaire':Titulaire,
                'Min':Min,
                'Mindivby90':Mindivby90,
                'Buts':Buts,
                'PD':PD,
                'B_PD':B_PD,
                'B_PD':B_PD,
                'B_PenM':B_PenM,
                'PenM':PenM,
                'PenT':PenT,
                'CJ':CJ,
                'CR':CR,
                'xG':xG,
                'npxG':npxG,
                'xAG':xAG,
                'npxG_xAG':npxG_xAG,
                'PrgC':PrgC,
                'PrgP':PrgP,
                'PrgR':PrgR,
                'Butsby90':Butsby90,
                'PDby90':PDby90,
                'B_PDby90':B_PDby90,
                'BPenMby90':BPenMby90,
                'B_PD_PenMby90':B_PD_PenMby90,
            })
    return render(request, 'pages/playerstats.html', {'data':data})

@login_required(login_url='loginpage')
def get_pronostic(request):  
    
    data_matches = openpyxl.load_workbook('data_matches.xlsx')
    sheet = data_matches['features']
    today = date.today()
    seven_days_after = today + timedelta(days=7)
    
    team_images = {
        'Real Madrid' : 'team_images/real-madrid.png',  
        'FC Barcelona' : 'team_images/fc-barcelona.png',
        'CA Osasuna' : 'team_images/osasuna.png',
        'RC Celta' : 'team_images/celta.png',
        'Real Valladolid CF' : 'team_images/valladolid.png',
        'Cádiz CF' : 'team_images/cadiz.png',  
        'Valencia CF' : 'team_images/valencia.png',  
        'UD Almería' : 'team_images/almeria.png',  
        'Athletic Club' : 'team_images/athletic.png',  
        'Getafe CF' : 'team_images/getafe.png',  
        'Real Betis' : 'team_images/real-betis.png',  
        'Sevilla FC' : 'team_images/sevilla.png',  
        'RCD Espanyol' : 'team_images/espanyol.png',  
        'Villarreal CF' : 'team_images/villarreal.png',  
        'Rayo Vallecano' : 'team_images/rayo.png',  
        'Real Sociedad' : 'team_images/real-sociedad.png',  
        'Girona FC' : 'team_images/girona.png',  
        'RCD Mallorca' : 'team_images/mallorca.png',  
        'Atlético de Madrid' : 'team_images/atletico-de-madrid.png',  
        'Elche CF' : 'team_images/elche.png',  
    }
    team_codes = {
    'Real Madrid': 33,
    'FC Barcelona': 5,
    'CA Osasuna': 31,
    'RC Celta': 8,
    'Real Valladolid CF': 42,
    'Cádiz CF': 7,
    'Valencia CF': 41,
    'UD Almería': 2,
    'Athletic Club': 3,
    'Getafe CF': 15,
    'Real Betis': 6,
    'Sevilla FC': 37,
    'RCD Espanyol': 13,
    'Villarreal CF': 45,
    'Rayo Vallecano': 43,
    'Real Sociedad': 38,
    'Girona FC': 17,
    'RCD Mallorca': 27,
    'Atlético de Madrid': 4,
    'Elche CF': 12
}


    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        home_team = row[9]
        away_team = row[12]
        match_time = row[2]
        stadium = row[8]
        match_date = row[2].date()
        
        # Convert dates to month/day tuples for comparison
        match_month_day = (match_date.month, match_date.day)
        today_month_day = (today.month, today.day)
        seven_days_after_month_day = (seven_days_after.month, seven_days_after.day)

        if match_month_day > seven_days_after_month_day:
            break
        elif match_month_day >= today_month_day:
            data.append({
                'home_team': home_team,
                'away_team': away_team,
                'match_time': match_time,
                'stadium': stadium,
                'home_team_images': team_images[home_team],
                'away_team_images': team_images[away_team],
                'ht_code': team_codes[home_team],
                'at_code': team_codes[away_team],
            })

    return render(request,'pages/get_pronostic.html',{'data':data})

model_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'machinemodel.pkl')
with open(model_path, 'rb') as f:
     machinemodel = pickle.load(f)
def get_predictions(request):
    # get the home and away team codes from the GET request
    ht_code = request.GET.get('ht_code', None)
    at_code = request.GET.get('at_code', None)

    # predict the winner using the loaded model and team codes
    prediction = machinemodel.predict([[ht_code, at_code]])

    # prepare the response as a dictionary
    response_data = {
        'prediction': prediction.tolist()[0]
    }

    # return the response as JSON
    return JsonResponse(response_data)

def loginpage(request):
    if request.user.is_authenticated:
        return redirect('home')
    else:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')

            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                messages.info(request, 'Username OR Password is incorrect !!!')
        context={}
        return render(request, 'pages/loginpage.html', context)

def logoutUser(request):
    logout(request)
    return redirect('home')

def registerpage(request):
    if request.user.is_authenticated:
        return redirect('home')
    else:
        form = CreateUserForm()
        if request.method == 'POST':
            form = CreateUserForm(request.POST)
            if form.is_valid():
                form.save()
                user = form.cleaned_data.get('username')
                messages.success(request, 'Account has been created for '+ user)
                return redirect('loginpage')
        context = {'form':form}
        return render(request, 'pages/registerpage.html', context)

@login_required(login_url='loginpage')
def userprofil(request):
    if request.method == 'POST':
        form = UpdateUserForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile was updated successfully!')
            return redirect('userprofil')
    else:
        form = UpdateUserForm(instance=request.user)
    context = {'form':form}
    return render(request, 'pages/userprofil.html', context)

